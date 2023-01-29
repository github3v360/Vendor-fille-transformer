from openpyxl import load_workbook
import openpyxl
import os
import pandas as pd
from src.utils import data_cleaner
import re
 
 
def Find(string):
 
    # Use a regular expression pattern to match URLs
    pattern = re.compile(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+')
    links = re.findall(pattern, string)
    
    # Return the first link if any are found, or None if not
    return links[0] if links else None

test_data_dir  = os.path.join("artifacts","new_test_data")
test_data_files_name =  os.listdir(test_data_dir)

import pandas as pd
from openpyxl import load_workbook

for file_name in test_data_files_name[:]:
    # if file_name != "JBBROTHER.xlsx":
    #     continue
    print()
    print(f" =========== File Name : {file_name} ============")
    df = pd.read_excel(os.path.join(test_data_dir,file_name), None)
    sheet_names = list(df.keys())
    cur_sheet = df[sheet_names[0]]
    columns_name = list(cur_sheet.columns)

    wb = openpyxl.load_workbook(os.path.join(test_data_dir,file_name))
    ws = wb[sheet_names[0]]

    cols_link = []

    for cur_col in range(1,len(columns_name)+1):
        
        if ws.cell(row = 12, column = cur_col).value is not None:
            cur_link = ws.cell(row = 12, column = cur_col).hyperlink

            if cur_link is not None:
                if cur_link.target is not None:
                    cols_link.append((columns_name[cur_col-1],cur_col))

            else:
                if type(ws.cell(row = 12, column = cur_col).value) == str:
                    if Find(ws.cell(row = 12, column = cur_col).value) is not None:
                        cols_link.append((columns_name[cur_col-1],cur_col))


    
    print(len(cols_link))
    


    # cur_sheet = df[sheet_names[0]]
    # df,correct_row_idx = data_cleaner.correct_df_headers(cur_sheet)
    # # print(correct_row_idx)
    # # print(len(df))
    # # print(ws.cell(row = correct_row_idx+2, column = cols_link[0][1]).hyperlink)
    # # print(ws.cell(row = correct_row_idx+3, column = cols_link[0][1]).hyperlink)
    # # print(df.columns[cols_link[0][1]-1])
    
    # df_link = pd.DataFrame()
    # for j in range(len(cols_link)):
    #     df_link[df.columns[cols_link[j][1]-1] + "_link"] = [None]*len(df)


    # t = 0
    # for i in range(correct_row_idx+3,ws.max_row+1):

    #     for j in range(len(cols_link)):
    #         if ws.cell(row=i,column=cols_link[j][1]).hyperlink is None:
    #             df_link[df.columns[cols_link[j][1]-1] + "_link"].iloc[t] = None
    #             continue
    #         df_link[df.columns[cols_link[j][1]-1] + "_link"].iloc[t] = ws.cell(row=i,column=cols_link[j][1]).hyperlink.target
    #     t+=1
    # print(df_link.head(2))
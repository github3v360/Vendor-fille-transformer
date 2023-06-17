import unittest
import pandas as pd
from openpyxl import Workbook
from src.hyperlink_extraction import HyperlinkExtractor
from openpyxl import load_workbook
from src.utils import data_cleaner
from src.utils import hyperlink_extraction_utils
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

class TestHyperlinkExtractor(unittest.TestCase):

    def setUp(self):

        '''
        We are going to use the hyperlink_testing_file.xlsx file which contains different 
        types of link for our testing. See the file to get the proper insight
        '''
        self.ws = load_workbook('artifacts/files_for_unit_testing/hyperlink_testing_file.xlsx')['Sheet1']
        self.df,self.correct_row_idx = data_cleaner.correct_df_headers(pd.read_excel('artifacts/files_for_unit_testing/hyperlink_testing_file.xlsx', engine='openpyxl'))

        self.hyperlink_extractor = HyperlinkExtractor(self.ws, self.correct_row_idx, self.df)
    
    def test_all_functions_in_hyperlink_extraction(self):

        '''
        ======= Test 1: def add_hyperlink_columns() testing ===========
        
        Input data frame had (5 rows and 13 columns)
        Since input dataframe had 2 link and in on of the link there is report number too
        Therefore 3 columns will be extracted 2 link columns 
        and 1 column with report number which was extracted from the link
        '''
        df, new_columns = self.hyperlink_extractor.add_hyperlink_columns()
        expected_new_columns =  ['Image_link', 'Cert._link', 'report_no']
        self.assertEqual(df.shape, (5, 16))
        self.assertEqual(new_columns,expected_new_columns)
        print("============ Test 1: def add_hyperlink_columns() testing successful ==========")

        '''
        ======= Test 2: def extract_link_from_current_cell() testing ===========
        Here we test whether we are able to extract the link from one of the 
        cell in excel sheet 

        We are taking cell at row = 6 and column = 13
        The cell value looks like this = " =HYPERLINK("http://online.sheetalgroup.com/certificate/1425013985.pdf","1425013985") "
        '''
        cur_cell = self.ws.cell(row=6, column=13)
        extracted_link = self.hyperlink_extractor.extract_link_from_current_cell(cur_cell)
        self.assertEqual(extracted_link,"http://online.sheetalgroup.com/certificate/1425013985.pdf")
        print("============ Test 2: def extract_link_from_current_cell() successful ==========")
        

        '''
        ======= Test 3: def iteratively_extract_link_and_report_number() testing ===========

        Here we test whether we are able to extract the link and report number from every cell by going to each cell iteratively
        Output would be the the dataframe with links and report number
        '''
        df_link = self.hyperlink_extractor.iteratively_extract_link_and_report_number(0)
        expected_columns =  ['Image_link', 'Cert._link', 'report_no']

        self.assertEqual(df_link.shape, (5, 3))
        self.assertEqual(list(df_link.columns),expected_columns)

        expected_extracted_report_no = ['514268','1425013985',None,'6431265811',None]
        self.assertEqual(expected_extracted_report_no,list(df_link['report_no']))

        expected_image_link_for_first_row = "https://dnadiamond.net/index.html?Bit=DB&Frm=Exc&refno=1.00W825086"
        expected_cert_link_for_second_row = "http://online.sheetalgroup.com/certificate/1425013985.pdf"
        self.assertEqual(df_link.Image_link.iloc[0],expected_image_link_for_first_row)
        self.assertEqual(df_link['Cert._link'].iloc[1],expected_cert_link_for_second_row)

        print("============ Test 3: def iteratively_extract_link_and_report_number() testing  successful ==========")

if __name__ == '__main__':
    unittest.main()
